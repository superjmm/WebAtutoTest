#coding=utf-8
import openpyxl

class HandleExcel:
    def load_excel(self,file_path):
        '''
        加载Excel文件
        param:
            file_path:excel文件路径+文件名
        '''
        try:
            wb = openpyxl.load_workbook(file_path)
        except:
            wb = None
            print("读取"+file_path+"文件出错！")
        return wb
    def get_sheet(self,file_path,sheet_num=None):
        '''
        拿到sheet对象
        param:
            file_path:excel文件路径+文件名
            sheet_num:sheet号，从0开始
        '''
        wb = self.load_excel(file_path)
        sheet = None
        if wb is not None:
            sheet_names = wb.sheetnames
            if sheet_num is None:
                sheet_num = 0
            sheet = wb[sheet_names[sheet_num]]
        return sheet

    def get_rows_num(self,file_path,sheet_num=None):
        '''
        拿到行数
        param:
            file_path:excel文件路径+文件名
            sheet_num:sheet号，从0开始
        '''
        sheet = self.get_sheet(file_path,sheet_num)
        row_num = 0
        if sheet is not None:
            row_num = sheet.max_row
        return row_num

    def get_row_value(self,file_path,sheet_num=None,row_num=None):
        '''
        拿到某一行的内容
        param:
            file_path:excel文件路径+文件名
            sheet_num:sheet号
            row_num:行号,表中是第几行，数字就是几

        '''
        if row_num is None:
            row_num = 0
        value_list = []
        sheet = self.get_sheet(file_path,sheet_num)
        if sheet is not None:
            for n in sheet[row_num]:
                value_list.append(n.value)
        return value_list
    
    def get_cell_value(self,file_path,sheet_num=None,row_num=None,col_num=None):
        '''
        拿到单元格内容
        param:
            file_path:excel文件路径+文件名
            sheet_num:sheet号
            row_num:行号,表中是第几行，数字就是几
            col_num:列号,表中是第几列，数字就是几
        '''
        sheet = self.get_sheet(file_path,sheet_num)
        cell_value = ''
        if sheet is not None:
            cell_value = sheet.cell(row_num,col_num).value
        return cell_value
    
    def get_table_value(self,file_path,sheet_num=None):
        '''
        拿到表格全部内容
        param:
            file_path:excel文件路径+文件名
            sheet_num:sheet号
            row_num:行号
        '''
        value_list = []
        rows_num = self.get_rows_num(file_path)
        if rows_num !=0:
            for n in range(2,rows_num+1):
                value_list.append(self.get_row_value(file_path,row_num=n))
        return value_list
    
    def write_value(self,file_path,row_num,col_num,value):
        '''
        写数据到某个单元格
        param:
            file_path:excel文件路径+文件名
            row_num:行号
            col_num:列号
            value:值
        '''
        wb = self.load_excel(file_path)
        wr = wb.active
        try:
            wr.cell(row_num,col_num,value)
            wb.save(file_path)
        except:
            print("写入"+file_path+"文件出错！")
     
handle_excel = HandleExcel()